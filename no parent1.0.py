import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# 全局变量，用于存储用户输入和 ASIN 文件
shop_name = ""
name = ""
brand = ""
product = ""
bottle_num = ""
asin_file = ""
asin_df = None  # 用于存储 ASIN 文件中的数据
asin_count = 0  # 用于存储 ASIN 的数量
datetime_str = datetime.now().strftime('%m%d')  # 当前日期格式为 'MMDD'
variation_theme = ""  # 全局变量用于存储变体主题

# 获取用户输入
def get_user_input():
    global name, brand, product, bottle_num, asin_file, asin_df, asin_count, variation_theme

    name = input("请输入名字(CZS): ").upper()
    brand = input("请输入品牌(HM、RR): ").upper()
    product = input("请输入产品(HG、HO、HOP、NK、MGG、TCG): ").upper()

    # 获取瓶装数
    if product in ['HG', 'HO', 'HOP','MGG','TCG']:
        while True:
            bottle_num = input("请输入瓶装数(数字): ")
            if bottle_num.isdigit():
                bottle_num += 'P'
                break
            else:
                print("输入无效，请输入数字！")

    # 用户选择变体主题
    print("请选择变体主题：")
    print("1. Flavor")
    print("2. SizeName")
    print("3. ColorName")
    print("4. SizeName-ColorName")
    print("5. Flavor-Size")
    print("6. Color")
    # print("7. Style")

    variation_theme_choices = {
        "1": "Flavor",
        "2": "SizeName",
        "3": "ColorName",
        "4": "SizeName-ColorName",
        "5": "Flavor-Size",
        "6": "Color",
        "7": "style"
    }

    while True:
        variation_theme_choice = input("请输入数字选择变体主题 (1-7): ")
        if variation_theme_choice in variation_theme_choices:
            variation_theme = variation_theme_choices[variation_theme_choice]
            break
        else:
            print("输入无效，请选择1到7之间的数字！")

    # 允许用户从本地选择 ASIN 文件（只执行一次）
    Tk().withdraw()  # 隐藏 Tkinter 主窗口
    asin_file = askopenfilename(title="请选择 ASIN 文件", filetypes=[("Excel files", "*.xlsx")])

    # 读取 ASIN 文件
    asin_df = pd.read_excel(asin_file)
    asin_count = asin_df['ASIN'].count()  # 根据实际的列名替换 'ASIN'

# 表1的处理函数
def process_table_1(target_file, start_row=4):
    global name, brand, product, bottle_num, asin_count, datetime_str

    # 读取目标文件
    wb = load_workbook(target_file)
    sheet = wb.active  # 选择活动工作表

    # 清空A4到X列的现有数据
    for row in range(start_row, start_row + asin_count):
        for col in ['B', 'D', 'E', 'Y']:  # 清空B、D、E、Y列
            cell = sheet[f'{col}{row}']
            cell.value = None  # 清空单元格内容

    # 插入数据到B列、D列、E列、Y列
    for i in range(asin_count):
        cell_b = f'B{start_row + i}'
        cell_e = f'E{start_row + i}'
        cell_y = f'Y{start_row + i}'

        # 插入数据到B列
        sequence_number = i + 1
        if bottle_num:
            spawn_sku = f'{brand}-{product}-{bottle_num}-{name}-{datetime_str}-{sequence_number}'
        else:
            spawn_sku = f'{brand}-{product}-{name}-{datetime_str}-{sequence_number}'

        sheet[cell_b] = spawn_sku

        # 插入数据到D列
        if 'ASIN' in asin_df.columns:
                asin_column_index = asin_df.columns.get_loc('ASIN')  # 获取 'ASIN' 列的索引
                asin_data = asin_df.iloc[:, asin_column_index].dropna().tolist()  # 获取该列下面的所有数据
                for i, value in enumerate(asin_data):
                    sheet[f'D{start_row + i}'] = value

        # 插入数据到E、Y列
        for i in range(start_row, start_row + asin_count):
            cell_e = f'E{i}'
            cell_y = f'Y{i}'
            sheet[cell_e] = 'ASIN'
            sheet[cell_y] = 'Update'

    # 另存为新文件，不替换模板
    if bottle_num:
        new_filename = f'GM-{brand}{product}{bottle_num}-{name}-{datetime_str}.xlsx'
    else:
        new_filename = f'GM-{brand}{product}-{name}-{datetime_str}.xlsx'
    wb.save(f'./无父体1.0合并表/{new_filename}')

    print(f"跟卖表处理完成: {new_filename}")

# 表2的处理函数
def process_table_2(target_file, start_row=4):
    global name, brand, product, bottle_num, asin_file, asin_df, asin_count, variation_theme

    # 读取目标文件
    wb = load_workbook(target_file)
    sheet = wb.active  # 选择活动工作表

    # 清空 B4 到 X 列的现有数据
    for row in range(start_row, start_row + asin_count + 1):  # +1 for the additional row
        for col in [ 'B', 'D', 'I', 'J', 'AJ', 'AK', 'AL', 'AM', 'AW', 'AX', 'AY', 'BB', 'BD']:  # 清空列
            cell = sheet[f'{col}{row}']
            cell.value = None  # 清空单元格内容

    # 插入数据到 A 列
    cell_a = f'A{start_row}'
    if product in ['HG', 'HO','MGG','TCG']:
        sheet[cell_a] = 'nutritionalsupplement'
    elif product == 'HOP':
        sheet[cell_a] = 'petsuppliesmisc'
    elif product == 'NK':
        sheet[cell_a] = 'underpants'

    # 插入数据到 B 列
    if bottle_num:
        # Insert SKU without sequence number in B4
        cell_b = f'B{start_row}'
        if bottle_num:
            spawn_sku_no_seq = f'{brand}-{product}-{bottle_num}-{name}-{datetime_str}'
        else:
            spawn_sku_no_seq = f'{brand}-{product}-{name}-{datetime_str}'
        sheet[cell_b] = spawn_sku_no_seq

        # Insert SKU with sequence number from B5 onwards
        for i in range(asin_count):
            cell_b = f'B{start_row + i + 1}'
            sequence_number = i + 1
            spawn_sku = f'{spawn_sku_no_seq}-{sequence_number}'
            sheet[cell_b] = spawn_sku

    # 插入数据到 C 列
    cell_c = f'C{start_row}'
    if brand == 'CN':
        sheet[cell_c] = 'CHANUBITO'
    elif brand == 'HM':
        sheet[cell_c] = 'HMone'
    elif brand == 'RRH':
        sheet[cell_c] = 'RiRywony Health'
    elif brand == 'WA':
        sheet[cell_c] = 'WACHRAY'
    elif brand == 'NT':
        sheet[cell_c] = 'NUBETONG'
    elif brand == 'LH':
        sheet[cell_c] = 'letollhold'
    elif brand == 'LC':
        sheet[cell_c] = 'LUCKCHAN'
    elif brand == 'MX':
        sheet[cell_c] = 'MaxHemp'
    elif brand == 'DR':
        sheet[cell_c] = 'Drloton'
    elif brand == 'HEM':
        sheet[cell_c] = 'HEMOMAC'
    elif brand == 'SU':
        sheet[cell_c] = 'SUXHDRPURE'
    elif brand == 'TA':
        sheet[cell_c] = 'TOPCAPAK'
    elif brand == 'MO':
        sheet[cell_c] = 'MOSRAY'
    elif brand == 'HZ':
        sheet[cell_c] = 'Hoozzch'
    elif brand == 'HY':
        sheet[cell_c] = 'Hemyum'
    elif brand == 'ZG':
        sheet[cell_c] = 'ziehooGe'
    elif brand == 'NR':
        sheet[cell_c] = 'NLMUBR LUCKSIT'
    elif brand == 'SUO':
        sheet[cell_c] = 'SUOOCH'
    elif brand == 'RR':
        sheet[cell_c] = 'RiRywony'
    elif brand == 'CF':
        sheet[cell_c] = 'CHICFAN'
    elif brand == 'MA':
        sheet[cell_c] = 'MAMOWYZ'

    # 插入数据到 D 列
    for i in range(start_row, start_row + asin_count + 1):
        cell_d = f'D{i}'
        if i == start_row:
            sheet[cell_d] = 'Update'
        else:
            sheet[cell_d] = 'PartialUpdate'

    # 插入数据到 I 列，从上传文件中的 'ASIN' 列中提取数据
    if 'ASIN' in asin_df.columns:
        asin_column_index = asin_df.columns.get_loc('ASIN')  # 获取 'ASIN' 列的索引
        asin_data = asin_df.iloc[:, asin_column_index].dropna().tolist()  # 获取该列下面的所有数据
        for i, value in enumerate(asin_data):
            sheet[f'I{start_row + i + 1}'] = value

    # 插入数据到 J 列
    for i in range(start_row, start_row + asin_count + 1):
        sheet[f'J{i}'] = 'ASIN'

    # 插入数据到 E 列
    header_name = '标题'
    # 检查标题列是否存在
    if header_name in asin_df.columns:
        title_column_index = asin_df.columns.get_loc(header_name)  # 获取标题列的索引
        title_value = asin_df.iloc[0, title_column_index]  # 获取标题下面的第一个数据
        sheet[f'E{start_row}'] = title_value
    else:
        print(f"警告: '{header_name}' 列不存在于 ASIN 文件中")

    # 插入数据到 AJ 列
    for i in range(start_row + 1, start_row + asin_count + 1):
        cell_aj = f'AJ{i}'
        sheet[cell_aj].value = sheet[f'B{start_row}'].value  # Copy value from B4
        cell_al = f'AL{i}'
        if i > start_row:
            sheet[cell_al] = 'Variation'

    # 插入数据到 AK 列
    for i in range(start_row, start_row + asin_count + 1):
        cell_ak = f'AK{i}'
        if i == start_row:
            sheet[cell_ak] = 'Parent'
        else:
            sheet[cell_ak] = 'Child'
        sheet[f'AM{i}'] = variation_theme

    # 根据 variation_theme 插入数据到 AW, AX, AY, BD 列
    if variation_theme == "Flavor":
        flavor_data = asin_df['Flavor'].dropna().tolist()
        for i, value in enumerate(flavor_data):
            sheet[f'AW{start_row + i + 1}'] = value

    elif variation_theme == "SizeName":
        size_data = asin_df['Keepa_Size'].dropna().tolist()
        for i, value in enumerate(size_data):
            sheet[f'AX{start_row + i + 1}'] = value

    elif variation_theme == "ColorName":
        color_data = asin_df['Keepa_Color'].dropna().tolist()
        for i, value in enumerate(color_data):
            sheet[f'AY{start_row + i + 1}'] = value
            sheet[f'BD{start_row + i + 1}'] = value

    elif variation_theme == "SizeName-ColorName":
        size_data = asin_df['Keepa_Size'].dropna().tolist()
        color_data = asin_df['Keepa_Color'].dropna().tolist()
        for i, (size, color) in enumerate(zip(size_data, color_data)):
            sheet[f'AX{start_row + i + 1}'] = size
            sheet[f'AY{start_row + i + 1}'] = color
            sheet[f'BD{start_row + i + 1}'] = color

    elif variation_theme == "Flavor-Size":
        flavor_data = asin_df['Flavor'].dropna().tolist()
        size_data = asin_df['Keepa_Size'].dropna().tolist()
        for i, (flavor, size) in enumerate(zip(flavor_data, size_data)):
            sheet[f'AW{start_row + i + 1}'] = flavor
            sheet[f'AX{start_row + i + 1}'] = size

    elif variation_theme == "Color":
        color_data = asin_df['Keepa_Color'].dropna().tolist()
        for i, value in enumerate(color_data):
            sheet[f'AY{start_row + i + 1}'] = value
            sheet[f'BD{start_row + i + 1}'] = value

    # 另存为新文件，不替换模板
    if bottle_num:
        new_filename = f'{brand}{product}{bottle_num}-{name}-{datetime_str}.xlsx'
    else:
        new_filename = f'{brand}{product}-{name}-{datetime_str}.xlsx'
    wb.save(f'./无父体1.0合并表./{new_filename}')

    print(f"合并表处理完成: {new_filename}")


# 调用全局输入和文件读取
get_user_input()

process_table_1('跟卖表.xlsx')
process_table_2('合并表.xlsx')