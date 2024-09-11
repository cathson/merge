import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# 全局变量，用于存储用户输入和 ASIN 文件
shop_name = ""
name = ""
brand = ""
product = ""
bottle_num = ""
asin_file = ""
asin_df = None  # 用于存储 ASIN 文件中的数据
asin_count = 0  # 用于存储 ASIN 的数量
datetime_str = datetime.now().strftime('%m%d')  # 当前日期格式为 'MMDD'
variation_theme = ""  # 全局变量用于存储变体主题
Parent_SKU = "" # 用来存储父SKU

# 获取用户输入
def get_user_input():
    global shop_name, name, brand, product, bottle_num, asin_file, asin_df, asin_count, variation_theme,Parent_SKU

    shop_name = input("请输入店铺名(SB、T29): ").upper()
    name = input("请输入名字(CZS): ").upper()
    brand = input("请输入品牌(HM、RR): ").upper()
    product = input("请输入产品(HG、HO、HOP、NK、MGG): ").upper()

    # 获取瓶装数
    if product in ['HG', 'HO', 'HOP','MGG']:
        while True:
            bottle_num = input("请输入瓶装数(数字): ")
            if bottle_num.isdigit():
                bottle_num += 'P'
                break
            else:
                print("输入无效，请输入数字！")

    # 用户选择变体主题
    print("请选择变体主题：")
    print("1. Flavor")
    print("2. SizeName")
    print("3. ColorName")
    print("4. SizeName-ColorName")
    print("5. Flavor-Size")
    print("6. Color")
    # print("7. Style")

    variation_theme_choices = {
        "1": "Flavor",
        "2": "SizeName",
        "3": "ColorName",
        "4": "SizeName-ColorName",
        "5": "Flavor-Size",
        "6": "Color",
        "7": "style"
    }

    while True:
        variation_theme_choice = input("请输入数字选择变体主题 (1-7): ")
        if variation_theme_choice in variation_theme_choices:
            variation_theme = variation_theme_choices[variation_theme_choice]
            break
        else:
            print("输入无效，请选择1到7之间的数字！")

    Parent_SKU = input("请输入目标ASIN的父SKU: ")

    # 允许用户从本地选择 ASIN 文件（只执行一次）
    Tk().withdraw()  # 隐藏 Tkinter 主窗口
    asin_file = askopenfilename(title="请选择 ASIN 文件", filetypes=[("Excel files", "*.xlsx")])

    # 读取 ASIN 文件
    asin_df = pd.read_excel(asin_file)
    asin_count = asin_df['ASIN'].count()  # 根据实际的列名替换 'ASIN'

# 表1的处理函数
def process_table_1(target_file, start_row=4):
    global shop_name, name, brand, product, bottle_num, asin_count, datetime_str

    # 读取目标文件
    wb = load_workbook(target_file)
    sheet = wb.active  # 选择活动工作表

    # 清空A4到X列的现有数据
    for row in range(start_row, start_row + asin_count):
        for col in ['A', 'B', 'F', 'X']:  # 清空A、B、F、X列
            cell = sheet[f'{col}{row}']
            cell.value = None  # 清空单元格内容

    # 插入数据到A列、B列、F列、X列
    for i in range(asin_count - 1):
        cell_a = f'A{start_row + i}'
        cell_b = f'B{start_row + i}'
        cell_f = f'F{start_row + i}'
        cell_x = f'X{start_row + i}'

        # 插入数据到A列
        sheet[cell_a] = 'comicbook'

        # 插入数据到B列
        sequence_number = i + 1
        if bottle_num:
            spawn_sku = f'{shop_name}-{brand}{product}{bottle_num}-{name}-{datetime_str}-{sequence_number}'
        else:
            spawn_sku = f'{shop_name}-{brand}{product}-{name}-{datetime_str}-{sequence_number}'

        sheet[cell_b] = spawn_sku

        # 插入数据到F列
        sheet[cell_f] = 1

        # 插入数据到X列
        sheet[cell_x] = 'PartialUpdate'

    # 另存为新文件，不替换模板
    new_filename = f'{brand}{product}{bottle_num}-{name}-{datetime_str}-1.xlsx'
    wb.save(f'./有父体2.0合并表/{new_filename}')

    print(f"表1处理完成: {new_filename}")

# 表2的处理函数
def process_table_2(target_file, start_row=4):
    global shop_name, name, brand, product, bottle_num, asin_file, asin_df, asin_count, variation_theme

    # 读取目标文件
    wb = load_workbook(target_file)
    sheet = wb.active  # 选择活动工作表

    # 清空 A4 到 X 列的现有数据
    for row in range(start_row, start_row + asin_count + 1):  # +1 for the additional row
        for col in ['A', 'B', 'D', 'E', 'J', 'AJ', 'AK', 'AL', 'AM', 'AW', 'AX', 'AY', 'BD']:  # 清空列
            cell = sheet[f'{col}{row}']
            cell.value = None  # 清空单元格内容

    # 插入数据到 A 列
    for i in range(asin_count - 1):  # +1 for the additional row
        cell_a = f'A{start_row + i}'
        if product in ['HG', 'HO','MGG']:
            sheet[cell_a] = 'nutritionalsupplement'
        elif product == 'HOP':
            sheet[cell_a] = 'petsuppliesmisc'
        elif product == 'NK':
            sheet[cell_a] = 'underpants'

    # 插入数据到 B 列
    if bottle_num:
        cell_b = f'B{start_row}'
        for i in range(asin_count - 1):
            cell_b = f'B{start_row + i}'
            sequence_number = i + 1
            spawn_sku = f'{shop_name}-{brand}{product}{bottle_num}-{name}-{datetime_str}-{sequence_number}'
            sheet[cell_b] = spawn_sku

    # 插入数据到 D 列
    for i in range(start_row, start_row + asin_count - 1):
        cell_d = f'D{i}'
        sheet[cell_d] = 'PartialUpdate'
        sheet[f'J{i}'] = 'ASIN'
        cell_ak = f'AK{i}'
        sheet[cell_ak] = 'Child'
        cell_al = f'AL{i}'
        sheet[cell_al] = 'Variation'
        sheet[f'AM{i}'] = variation_theme
        cell_aj = f'AJ{i}'
        sheet[cell_aj] = Parent_SKU

    # 根据 variation_theme 插入数据到 AW, AX, AY, BD 列
    if variation_theme == "Flavor":
        flavor_data = asin_df['Flavor'].dropna().tolist()[1:]
        for i, value in enumerate(flavor_data):
            sheet[f'AW{start_row + i}'] = value

    elif variation_theme == "SizeName":
        size_data = asin_df['Keepa_Size'].dropna().tolist()[1:]
        for i, value in enumerate(size_data):
            sheet[f'AX{start_row + i}'] = value

    elif variation_theme == "ColorName":
        color_data = asin_df['Keepa_Color'].dropna().tolist()[1:]
        for i, value in enumerate(color_data):
            sheet[f'AY{start_row + i}'] = value
            sheet[f'BD{start_row + i}'] = value

    elif variation_theme == "SizeName-ColorName":
        size_data = asin_df['Keepa_Size'].dropna().tolist()[1:]
        color_data = asin_df['Keepa_Color'].dropna().tolist()[1:]
        for i, (size, color) in enumerate(zip(size_data, color_data)):
            sheet[f'AX{start_row + i}'] = size
            sheet[f'AY{start_row + i}'] = color
            sheet[f'BD{start_row + i}'] = color

    elif variation_theme == "Flavor-Size":
        flavor_data = asin_df['Flavor'].dropna().tolist()[1:]
        size_data = asin_df['Keepa_Size'].dropna().tolist()[1:]
        for i, (flavor, size) in enumerate(zip(flavor_data, size_data)):
            sheet[f'AW{start_row + i}'] = flavor
            sheet[f'AX{start_row + i}'] = size

    elif variation_theme == "Color":
        color_data = asin_df['Keepa_Color'].dropna().tolist()[1:]
        for i, value in enumerate(color_data):
            sheet[f'AY{start_row + i}'] = value
            sheet[f'BD{start_row + i}'] = value

    # 另存为新文件，不替换模板
    new_filename = f'{brand}{product}{bottle_num}-{name}-{datetime_str}-2.xlsx'
    wb.save(f'./有父体2.0合并表./{new_filename}')

    print(f"表2处理完成: {new_filename}")

def process_table_3(target_file, start_row=4):
    global asin_df, asin_count, variation_theme

    # 读取表3目标文件
    wb_table3 = load_workbook(target_file)
    sheet = wb_table3.active

    # 清空现有数据（B列、D列、E列、F列、G列、H列、I列、J列、S列、FJ列）
    for row in range(start_row, start_row + asin_count):
        for col in ['B', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'S', 'FJ']:
            sheet[f'{col}{row}'].value = None

    # 插入数据到B列
    for i in range(asin_count - 1):
        cell_b = f'B{start_row + i}'

        # 生成 spawn_sku
        sequence_number = i + 1
        if bottle_num:
            spawn_sku = f'{shop_name}-{brand}{product}{bottle_num}-{name}-{datetime_str}-{sequence_number}'
        else:
            spawn_sku = f'{shop_name}-{brand}{product}-{name}-{datetime_str}-{sequence_number}'

        sheet[cell_b] = spawn_sku

    # 插入数据到 D 列，全部填入 'update'
    for i in range(asin_count - 1):
        sheet[f'D{start_row + i}'] = 'update'

    # 插入数据到 E 列，从上传文件中的 '标题' 列中提取数据
    if '标题' in asin_df.columns:
        title_column_index = asin_df.columns.get_loc('标题')  # 获取 '标题' 列的索引
        title_data = asin_df.iloc[:, title_column_index].dropna().tolist()[1:]  # 获取该列下面的所有数据
        for i, value in enumerate(title_data):
            sheet[f'E{start_row + i}'] = value

    # 插入数据到 F 列，从上传文件中的 'ASIN' 列中提取数据
    if 'ASIN' in asin_df.columns:
        asin_column_index = asin_df.columns.get_loc('ASIN')  # 获取 'ASIN' 列的索引
        asin_data = asin_df.iloc[:, asin_column_index].dropna().tolist()[1:]  # 获取该列下面的所有数据
        for i, value in enumerate(asin_data):
            sheet[f'F{start_row + i}'] = value

    # 插入数据到 G 列，全部填入 'ASIN'
    for i in range(asin_count - 1):
        sheet[f'G{start_row + i}'] = 'ASIN'

    # 插入数据到 H 列、I 列、J 列、S 列，根据变体主题插入相应数据
    if variation_theme == "Flavor":
        if 'Flavor' in asin_df.columns:
            flavor_data = asin_df['Flavor'].dropna().tolist()[1:]
            for i, value in enumerate(flavor_data):
                sheet[f'H{start_row + i}'] = value

    if variation_theme == "SizeName":
        if 'Keepa_Size' in asin_df.columns:
            size_data = asin_df['Keepa_Size'].dropna().tolist()[1:]
            for i, value in enumerate(size_data):
                sheet[f'I{start_row + i}'] = value

    if variation_theme == "ColorName":
        if 'Keepa_Color' in asin_df.columns:
            color_data = asin_df['Keepa_Color'].dropna().tolist()[1:]
            for i, value in enumerate(color_data):
                sheet[f'J{start_row + i}'] = value
                sheet[f'S{start_row + i}'] = value  # S 列与 J 列相同

    if variation_theme == "SizeName-ColorName":
        if 'Keepa_Size' in asin_df.columns and 'Keepa_Color' in asin_df.columns:
            size_data = asin_df['Keepa_Size'].dropna().tolist()[1:]
            color_data = asin_df['Keepa_Color'].dropna().tolist()[1:]
            for i, (size, color) in enumerate(zip(size_data, color_data)):
                sheet[f'I{start_row + i}'] = size
                sheet[f'J{start_row + i}'] = color
                sheet[f'S{start_row + i}'] = color  # S 列与 J 列相同

    # 插入数据到 FJ 列，全部填入 '1'
    for i in range(asin_count - 1):
        sheet[f'FJ{start_row + i}'] = 1

    # 另存为新文件
    new_filename = f'{brand}{product}{bottle_num}-{name}-{datetime_str}-3.xlsx'
    wb_table3.save(f'./有父体2.0合并表/{new_filename}')

    print(f"表3处理完成: {new_filename}")




# 调用全局输入和文件读取
get_user_input()

process_table_1('表1.xlsx')
process_table_2('表2.xlsx')
process_table_3('表3.xlsx')
